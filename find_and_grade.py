import argparse
import csv
import json
import os
import random
import re
from urllib.parse import urlparse
import time
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from google_sheets_export import export_to_google_sheets
from main import MAX_WORKERS, PRIORITY_OUTPUT_FILE, grade_website

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.nchc.org.tw/api/interpreter",
    "https://overpass.openstreetmap.ru/api/interpreter",
]
USER_AGENT = "WebsiteGrader-FindAndGrade/1.0 (contact: nathan.strickland.consult@gmail.com)"
NOMINATIM_EMAIL = "nathan.strickland.consult@gmail.com"
YELP_SEARCH_URL = "https://api.yelp.com/v3/businesses/search"
BING_LOCAL_SEARCH_URL = "https://api.bing.microsoft.com/v7.0/localbusinesses/search"
REQUEST_DELAY_SECONDS = 1.0
OVERPASS_MAX_RETRIES = 3
OVERPASS_BACKOFF_SECONDS = 2.0
OVERPASS_JITTER_SECONDS = 0.5
OVERPASS_RETRYABLE_STATUS_CODES = {429, 500, 502, 503, 504}
RAW_OUTPUT_FILE = "leads_raw.csv"
GRADED_OUTPUT_FILE = "report.csv"
PRIORITY_LEADS_OUTPUT_FILE = "priority_leads.csv"
OUTREACH_SHEET_OUTPUT_FILE = "outreach_sheet.xlsx"

WEBSITE_KEYS = ["website", "contact:website", "url"]
PHONE_KEYS = ["phone", "contact:phone"]
DEFAULT_TERMS = ["plumber", "plumbing", "drain", "sewer", "rooter", "pipe", "water heater"]
EXCLUDED_AMENITIES = {"school", "college", "university", "kindergarten"}
EXCLUDED_NAME_TERMS = ["school", "elementary", "middle school", "high school", "academy", "university"]
PLUMBING_NAME_KEYWORDS = ["plumb", "drain", "sewer", "water heater"]
DEFAULT_CATEGORY = "plumber"
CATEGORY_CONFIG: Dict[str, Dict[str, object]] = {
    "plumber": {
        "osm_pairs": [("shop", "plumber"), ("craft", "plumber")],
        "default_terms": DEFAULT_TERMS,
        "excluded_amenities": EXCLUDED_AMENITIES,
        "excluded_name_terms": EXCLUDED_NAME_TERMS,
        "name_keywords": PLUMBING_NAME_KEYWORDS,
    },
    "roofer": {
        "osm_pairs": [("craft", "roofer")],
        "default_terms": ["roofer", "roofing", "roof repair", "roof replacement"],
        "excluded_amenities": EXCLUDED_AMENITIES,
        "excluded_name_terms": EXCLUDED_NAME_TERMS,
        "name_keywords": ["roof", "roofer", "roofing"],
    },
    "hvac": {
        "osm_pairs": [("craft", "hvac"), ("shop", "hvac")],
        "default_terms": ["hvac", "heating", "cooling", "air conditioning", "furnace"],
        "excluded_amenities": EXCLUDED_AMENITIES,
        "excluded_name_terms": EXCLUDED_NAME_TERMS,
        "name_keywords": ["hvac", "heating", "cooling", "air conditioning", "furnace"],
    },
}


def clean_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def get_city_center(city: str) -> Tuple[float, float]:
    time.sleep(1)
    params = {"q": city, "format": "json", "limit": 1, "email": NOMINATIM_EMAIL}
    headers = {"User-Agent": USER_AGENT, "Accept-Language": "en"}
    response = requests.get(
        NOMINATIM_URL,
        params=params,
        headers=headers,
        timeout=20,
    )
    if response.status_code == 403:
        request_url = requests.Request("GET", NOMINATIM_URL, params=params).prepare().url
        print("Nominatim 403: check USER_AGENT and email param")
        print(f"Requested URL: {request_url}")
    response.raise_for_status()

    payload = response.json()
    if not payload:
        raise ValueError(f"Could not geocode city: {city}")

    lat = payload[0].get("lat")
    lon = payload[0].get("lon")
    if lat is None or lon is None:
        raise ValueError(f"City geocoding did not return a valid center point: {city}")

    return float(lat), float(lon)


def build_name_regex(terms: List[str]) -> str:
    escaped_terms = [re.escape(term.strip()) for term in terms if term.strip()]
    if not escaped_terms:
        escaped_terms = [re.escape(term) for term in DEFAULT_TERMS]
    return "|".join(escaped_terms)


def build_overpass_query(
    lat: float,
    lon: float,
    radius_km: float,
    osm_pairs: List[Tuple[str, str]],
    overpass_timeout: int,
) -> str:
    radius_m = int(radius_km * 1000)
    pair_clauses: List[str] = []
    for key, value in osm_pairs:
        escaped_key = str(key).strip()
        escaped_value = str(value).strip()
        if not escaped_key or not escaped_value:
            continue
        pair_clauses.extend(
            [
                f'  node["{escaped_key}"="{escaped_value}"](around:{radius_m},{lat},{lon});',
                f'  way["{escaped_key}"="{escaped_value}"](around:{radius_m},{lat},{lon});',
                f'  relation["{escaped_key}"="{escaped_value}"](around:{radius_m},{lat},{lon});',
            ]
        )

    if not pair_clauses:
        raise ValueError("No valid OSM tag pairs configured for selected category.")

    clauses_text = "\n".join(pair_clauses)
    return f"""
[out:json][timeout:{overpass_timeout}];
(
{clauses_text}
);
out center;
""".strip()


def build_overpass_name_fallback_query(
    lat: float,
    lon: float,
    radius_km: float,
    terms: List[str],
    overpass_timeout: int,
) -> str:
    radius_m = int(radius_km * 1000)
    name_regex = build_name_regex(terms)
    return f"""
[out:json][timeout:{overpass_timeout}];
(
  node["name"~"{name_regex}", i](around:{radius_m},{lat},{lon});
  way["name"~"{name_regex}", i](around:{radius_m},{lat},{lon});
  relation["name"~"{name_regex}", i](around:{radius_m},{lat},{lon});
);
out center;
""".strip()


def is_target_business(tags: Dict[str, str], name: str, category_config: Dict[str, object]) -> Tuple[bool, str]:
    normalized_tags = {str(k).strip().lower(): str(v).strip().lower() for k, v in (tags or {}).items()}
    normalized_name = (name or "").strip().lower()
    excluded_amenities = {
        str(amenity).strip().lower()
        for amenity in category_config.get("excluded_amenities", EXCLUDED_AMENITIES)
    }
    excluded_name_terms = [
        str(term).strip().lower()
        for term in category_config.get("excluded_name_terms", EXCLUDED_NAME_TERMS)
        if str(term).strip()
    ]
    osm_pairs = [
        (str(key).strip().lower(), str(value).strip().lower())
        for key, value in category_config.get("osm_pairs", [])
        if str(key).strip() and str(value).strip()
    ]
    name_keywords = [
        str(keyword).strip().lower()
        for keyword in category_config.get("name_keywords", [])
        if str(keyword).strip()
    ]

    amenity = normalized_tags.get("amenity", "")
    if amenity in excluded_amenities:
        return False, f"excluded_amenity={amenity}"

    for term in excluded_name_terms:
        if term in normalized_name:
            return False, f"excluded_name_term={term}"

    for key, value in osm_pairs:
        if normalized_tags.get(key) == value:
            return True, f"{key}={value}"

    for keyword in name_keywords:
        if keyword in normalized_name:
            return True, "name_keyword_fallback"

    return False, "missing_category_tag"


def fetch_overpass_elements(overpass_query: str, overpass_timeout: int) -> List[Dict[str, object]]:
    response_payload = call_overpass(overpass_query, overpass_timeout)
    return response_payload.get("elements", [])


def call_overpass(query: str, overpass_timeout: int) -> Dict[str, object]:
    last_exception: Optional[Exception] = None
    tried_endpoints: List[str] = []
    timeout_seconds = overpass_timeout + 30

    for endpoint in OVERPASS_ENDPOINTS:
        tried_endpoints.append(endpoint)

        for attempt in range(1, OVERPASS_MAX_RETRIES + 1):
            try:
                response = requests.post(
                    endpoint,
                    data={"data": query},
                    headers={"User-Agent": USER_AGENT},
                    timeout=timeout_seconds,
                )

                if response.status_code in OVERPASS_RETRYABLE_STATUS_CODES:
                    raise requests.HTTPError(
                        f"HTTP {response.status_code}",
                        response=response,
                    )
                response.raise_for_status()
                return response.json()
            except (requests.RequestException, ValueError) as exc:
                last_exception = exc
                error_details = str(exc)
                if isinstance(exc, requests.HTTPError) and exc.response is not None:
                    error_details = f"HTTP {exc.response.status_code}"

                print(
                    f"Overpass request failed at {endpoint} "
                    f"(attempt {attempt}/{OVERPASS_MAX_RETRIES}): {error_details}"
                )

                if attempt < OVERPASS_MAX_RETRIES:
                    backoff_seconds = OVERPASS_BACKOFF_SECONDS * (2 ** (attempt - 1))
                    jitter_seconds = random.uniform(0, OVERPASS_JITTER_SECONDS)
                    time.sleep(backoff_seconds + jitter_seconds)
                else:
                    print(f"Overpass endpoint exhausted: {endpoint}")

    endpoints_text = ", ".join(tried_endpoints)
    error_message = f"All Overpass endpoints failed after retries. Tried: {endpoints_text}"
    if last_exception is not None:
        raise RuntimeError(error_message) from last_exception
    raise RuntimeError(error_message)


def first_tag(tags: Dict[str, str], keys: List[str]) -> str:
    for key in keys:
        value = tags.get(key, "").strip()
        if value:
            return value
    return ""


def build_address(tags: Dict[str, str]) -> str:
    if tags.get("addr:full"):
        return clean_whitespace(tags["addr:full"])

    parts = [
        tags.get("addr:housenumber", "").strip(),
        tags.get("addr:street", "").strip(),
        tags.get("addr:city", "").strip(),
        tags.get("addr:state", "").strip(),
        tags.get("addr:postcode", "").strip(),
    ]
    parts = [part for part in parts if part]
    return clean_whitespace(", ".join(parts))


def parse_element(element: Dict[str, object], category: str, category_config: Dict[str, object]) -> Dict[str, object]:
    tags = element.get("tags", {}) or {}
    tags = {str(k): str(v) for k, v in tags.items()}
    include, inclusion_reason = is_target_business(tags, tags.get("name", ""), category_config)

    lat = element.get("lat")
    lon = element.get("lon")
    center = element.get("center") or {}
    if lat is None:
        lat = center.get("lat", "")
    if lon is None:
        lon = center.get("lon", "")

    osm_type = str(element.get("type", ""))
    osm_id = element.get("id", "")

    return {
        "name": tags.get("name", "").strip(),
        "business_name": tags.get("name", "").strip(),
        "website": first_tag(tags, WEBSITE_KEYS),
        "website_url": first_tag(tags, WEBSITE_KEYS),
        "phone": first_tag(tags, PHONE_KEYS),
        "address": build_address(tags),
        "lat": lat,
        "lon": lon,
        "osm_type": osm_type,
        "osm_id": osm_id,
        "osm_shop": tags.get("shop", ""),
        "osm_craft": tags.get("craft", ""),
        "osm_amenity": tags.get("amenity", ""),
        "osm_tags_json": json.dumps(tags, separators=(",", ":"), sort_keys=True),
        "is_target_business": include,
        "category": category,
        "inclusion_reason": inclusion_reason,
        "source": "osm",
        "sources": "osm",
        "source_listing_url": "",
        "source_id": f"{osm_type}/{osm_id}" if osm_type and osm_id else "",
    }


def normalize_domain(url: object) -> str:
    value = str(url or "").strip()
    if not value:
        return ""
    if not re.match(r"^https?://", value, flags=re.IGNORECASE):
        value = f"https://{value}"
    parsed = urlparse(value)
    domain = parsed.netloc.lower().strip()
    if domain.startswith("www."):
        domain = domain[4:]
    return domain


def normalize_phone(phone: object) -> str:
    return re.sub(r"\D", "", str(phone or ""))


def normalized_name_address(lead: Dict[str, object]) -> Tuple[str, str]:
    name = clean_whitespace(str(lead.get("business_name") or lead.get("name") or "").lower())
    address = clean_whitespace(str(lead.get("address", "")).lower())
    return name, address


def dedupe_leads(leads: List[Dict[str, object]]) -> List[Dict[str, object]]:
    deduped_map: Dict[str, Dict[str, object]] = {}

    for index, lead in enumerate(leads):
        website_domain = normalize_domain(lead.get("website_url") or lead.get("website"))
        phone = normalize_phone(lead.get("phone"))
        name, address = normalized_name_address(lead)

        if website_domain:
            dedupe_key = f"domain:{website_domain}"
        elif phone:
            dedupe_key = f"phone:{phone}"
        elif name and address:
            dedupe_key = f"name_address:{name}|{address}"
        else:
            dedupe_key = f"fallback:{index}"

        if dedupe_key not in deduped_map:
            lead_copy = dict(lead)
            if not str(lead_copy.get("business_name", "")).strip():
                lead_copy["business_name"] = str(lead_copy.get("name", "")).strip()
            if not str(lead_copy.get("name", "")).strip():
                lead_copy["name"] = str(lead_copy.get("business_name", "")).strip()
            if not str(lead_copy.get("website_url", "")).strip():
                lead_copy["website_url"] = str(lead_copy.get("website", "")).strip()
            if not str(lead_copy.get("website", "")).strip():
                lead_copy["website"] = str(lead_copy.get("website_url", "")).strip()
            source = str(lead_copy.get("source", "")).strip()
            lead_copy["sources"] = source
            deduped_map[dedupe_key] = lead_copy
            continue

        existing = deduped_map[dedupe_key]
        merged_sources = {
            source.strip()
            for source in f"{existing.get('sources', '')};{lead.get('source', '')}".split(";")
            if source.strip()
        }
        existing["sources"] = ";".join(sorted(merged_sources))

        for field in ("website", "website_url", "phone", "address", "source_listing_url", "source_id"):
            if not str(existing.get(field, "")).strip() and str(lead.get(field, "")).strip():
                existing[field] = lead.get(field, "")

    return list(deduped_map.values())


def sort_leads(leads: List[Dict[str, object]]) -> List[Dict[str, object]]:
    return sorted(
        leads,
        key=lambda lead: (
            str(lead.get("name", "")).strip().lower(),
            str(lead.get("address", "")).strip().lower(),
            str(lead.get("website", "")).strip().lower(),
            str(lead.get("osm_type", "")).strip().lower(),
            str(lead.get("osm_id", "")).strip(),
        ),
    )


def find_businesses(
    city: str,
    category: str,
    terms: List[str],
    limit: int,
    radius_km: float,
    overpass_timeout: int,
) -> List[Dict[str, object]]:
    return discover_leads_osm(
        city=city,
        category=category,
        limit=limit,
        radius_km=radius_km,
        terms=terms,
        overpass_timeout=overpass_timeout,
    )


def discover_leads_osm(
    city: str,
    category: str,
    limit: int,
    radius_km: float,
    terms: Optional[List[str]] = None,
    overpass_timeout: int = 180,
) -> List[Dict[str, object]]:
    category_config = CATEGORY_CONFIG[category]
    terms = terms or list(category_config.get("default_terms", DEFAULT_TERMS))
    lat, lon = get_city_center(city)
    time.sleep(REQUEST_DELAY_SECONDS)

    overpass_query = build_overpass_query(
        lat,
        lon,
        radius_km,
        list(category_config.get("osm_pairs", [])),
        overpass_timeout,
    )
    print("---- OVERPASS QUERY ----")
    print(overpass_query)
    print("------------------------")
    elements = fetch_overpass_elements(overpass_query, overpass_timeout)

    fallback_threshold = min(max(limit, 1), 10)
    if len(elements) < limit or len(elements) < fallback_threshold:
        fallback_query = build_overpass_name_fallback_query(
            lat=lat,
            lon=lon,
            radius_km=radius_km,
            terms=terms,
            overpass_timeout=overpass_timeout,
        )
        print("---- OVERPASS NAME FALLBACK QUERY ----")
        print(fallback_query)
        print("--------------------------------------")
        elements.extend(fetch_overpass_elements(fallback_query, overpass_timeout))

    leads = [
        lead
        for lead in (parse_element(element, category=category, category_config=category_config) for element in elements)
        if is_true(lead.get("is_target_business", False))
    ]
    ordered = sort_leads(leads)
    return ordered[:limit]


def discover_leads_yelp(city: str, limit: int, radius_km: float, category: str) -> List[Dict[str, object]]:
    api_key = os.getenv("YELP_API_KEY", "").strip()
    if not api_key:
        return []

    radius_m = min(max(int(radius_km * 1000), 1), 40000)
    params = {
        "term": category,
        "location": city,
        "limit": max(1, min(limit, 50)),
        "radius": radius_m,
    }
    headers = {"Authorization": f"Bearer {api_key}", "User-Agent": USER_AGENT}

    try:
        response = requests.get(YELP_SEARCH_URL, params=params, headers=headers, timeout=20)
        response.raise_for_status()
    except requests.RequestException as exc:
        print(f"Yelp discovery failed: {exc}")
        return []

    businesses = response.json().get("businesses", [])
    leads: List[Dict[str, object]] = []
    for business in businesses:
        location = business.get("location") or {}
        address_parts = [location.get("address1", ""), location.get("city", ""), location.get("state", ""), location.get("zip_code", "")]
        leads.append(
            {
                "name": str(business.get("name", "")).strip(),
                "business_name": str(business.get("name", "")).strip(),
                "website": "",
                "website_url": "",
                "phone": str(business.get("display_phone") or business.get("phone") or "").strip(),
                "address": clean_whitespace(", ".join(part for part in address_parts if part)),
                "lat": (business.get("coordinates") or {}).get("latitude", ""),
                "lon": (business.get("coordinates") or {}).get("longitude", ""),
                "source": "yelp",
                "category": category,
                "source_listing_url": str(business.get("url", "")).strip(),
                "source_id": str(business.get("id", "")).strip(),
            }
        )
    return leads


def discover_leads_bing(city: str, limit: int, radius_km: float, category: str) -> List[Dict[str, object]]:
    del radius_km
    api_key = os.getenv("BING_API_KEY", "").strip()
    if not api_key:
        return []

    params = {
        "q": f"{category} in {city}",
        "count": max(1, min(limit, 50)),
    }
    headers = {"Ocp-Apim-Subscription-Key": api_key, "User-Agent": USER_AGENT}

    try:
        response = requests.get(BING_LOCAL_SEARCH_URL, params=params, headers=headers, timeout=20)
        response.raise_for_status()
    except requests.RequestException as exc:
        print(f"Bing discovery failed: {exc}")
        return []

    values = response.json().get("value", [])
    leads: List[Dict[str, object]] = []
    for item in values:
        address_info = item.get("address") or {}
        address_parts = [address_info.get("addressLocality", ""), address_info.get("addressRegion", ""), address_info.get("postalCode", "")]
        leads.append(
            {
                "name": str(item.get("name", "")).strip(),
                "business_name": str(item.get("name", "")).strip(),
                "website": str(item.get("website", "")).strip(),
                "website_url": str(item.get("website", "")).strip(),
                "phone": str(item.get("telephone", "")).strip(),
                "address": clean_whitespace(", ".join(part for part in address_parts if part)),
                "lat": (item.get("geo") or {}).get("latitude", ""),
                "lon": (item.get("geo") or {}).get("longitude", ""),
                "source": "bing",
                "category": category,
                "source_listing_url": str(item.get("url", "")).strip(),
                "source_id": str(item.get("id", "")).strip(),
            }
        )
    return leads


def parse_lead_sources() -> List[str]:
    configured = os.getenv("LEAD_SOURCES", "osm")
    allowed = {"osm", "yelp", "bing"}
    sources = [source.strip().lower() for source in configured.split(",") if source.strip()]
    normalized = [source for source in sources if source in allowed]
    return normalized or ["osm"]


def parse_terms(terms_arg: str, default_terms: List[str]) -> List[str]:
    parsed = [term.strip() for term in terms_arg.split(",") if term.strip()]
    return parsed or default_terms.copy()


def write_raw_leads(path: str, leads: List[Dict[str, object]]) -> None:
    base_fieldnames = [
        "name",
        "business_name",
        "website",
        "website_url",
        "phone",
        "address",
        "lat",
        "lon",
        "source",
        "sources",
        "source_listing_url",
        "source_id",
        "osm_type",
        "osm_id",
        "osm_shop",
        "osm_craft",
        "osm_amenity",
        "osm_tags_json",
        "category",
        "inclusion_reason",
    ]
    extra_fieldnames = sorted(
        {
            str(key)
            for lead in leads
            for key in lead.keys()
            if str(key) not in base_fieldnames
        }
    )
    fieldnames = [*base_fieldnames, *extra_fieldnames]

    with open(path, "w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(leads)


def build_graded_row(lead: Dict[str, object], graded: Dict[str, object]) -> Dict[str, object]:
    return {
        "business_name": lead.get("business_name") or lead.get("name", ""),
        "phone": lead.get("phone", ""),
        "address": lead.get("address", ""),
        "lat": lead.get("lat", ""),
        "lon": lead.get("lon", ""),
        "source": lead.get("source", ""),
        "sources": lead.get("sources", lead.get("source", "")),
        "source_listing_url": lead.get("source_listing_url", ""),
        "source_id": lead.get("source_id", ""),
        "category": lead.get("category", ""),
        **graded,
    }


def write_graded_report(path: str, rows: List[Dict[str, object]]) -> None:
    ordered_fieldnames = [
        "business_name",
        "phone",
        "address",
        "lat",
        "lon",
        "source",
        "sources",
        "source_listing_url",
        "source_id",
        "category",
        "url",
        "reachable",
        "https",
        "status_code",
        "final_url",
        "title",
        "meta_description_present",
        "has_phone",
        "has_email",
        "primary_email",
        "emails",
        "has_contact_form",
        "contact_form_url",
        "outreach_channel",
        "form_message",
        "call_script",
        "has_contact_page_link",
        "contact_page_url",
        "contact_page_reachable",
        "notes",
        "reasons",
        "score_0_100",
    ]

    for row in rows:
        row.setdefault("opportunity_score_0_100", 0)
        row.setdefault("pitch", "")

    if not rows:
        fieldnames = [
            *ordered_fieldnames,
            "opportunity_score_0_100",
            "pitch",
        ]
    else:
        fieldnames = [
            *ordered_fieldnames,
            *[
                field
                for field in ("opportunity_score_0_100", "pitch")
                if field not in ordered_fieldnames
            ],
        ]

    with open(path, "w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def is_true(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() == "true"
    return False


def as_float(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            return 0.0
    return 0.0


def build_priority_row(row: Dict[str, object]) -> Dict[str, object]:
    return {
        "business_name": row.get("business_name", ""),
        "phone": row.get("phone", ""),
        "primary_email": row.get("primary_email", ""),
        "emails": row.get("emails", ""),
        "has_contact_form": row.get("has_contact_form", False),
        "contact_form_url": row.get("contact_form_url", ""),
        "outreach_channel": row.get("outreach_channel", "none"),
        "form_message": row.get("form_message", ""),
        "call_script": row.get("call_script", ""),
        "url": row.get("url", ""),
        "opportunity_score_0_100": row.get("opportunity_score_0_100", ""),
        "score_0_100": row.get("score_0_100", ""),
        "reasons": row.get("reasons", ""),
        "pitch": row.get("pitch", ""),
    }


def get_priority_rows(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    filtered_rows = [
        row
        for row in rows
        if is_true(row.get("reachable"))
        and is_true(row.get("has_phone"))
        and str(row.get("url", "")).strip()
        and as_float(row.get("opportunity_score_0_100", 0)) > 0
    ]
    return sorted(
        filtered_rows,
        key=lambda row: as_float(row.get("opportunity_score_0_100", 0)),
        reverse=True,
    )


def normalize_rows(rows: object) -> List[Dict[str, object]]:
    if isinstance(rows, dict):
        values = list(rows.values())
        if all(isinstance(value, dict) for value in values):
            return values
        example_value = values[0] if values else None
        raise TypeError(
            "write_priority_leads expected rows to be a list of dicts (or a dict of dicts); "
            f"got dict with value type {type(example_value).__name__}. Example value: {example_value!r}"
        )

    if isinstance(rows, list):
        if all(isinstance(row, dict) for row in rows):
            return rows
        example_row = rows[0] if rows else None
        raise TypeError(
            "write_priority_leads expected rows to be a list of dicts; "
            f"got list with element type {type(example_row).__name__}. Example element: {example_row!r}"
        )

    raise TypeError(
        "write_priority_leads expected rows to be a list of dicts (or a dict of dicts); "
        f"got {type(rows).__name__}. Example value: {rows!r}"
    )


def write_priority_leads(path: str, rows: List[Dict[str, object]]) -> int:
    normalized_rows = normalize_rows(rows)
    sorted_rows = get_priority_rows(normalized_rows)

    fieldnames = [
        "business_name",
        "phone",
        "primary_email",
        "emails",
        "has_contact_form",
        "contact_form_url",
        "outreach_channel",
        "form_message",
        "call_script",
        "url",
        "opportunity_score_0_100",
        "score_0_100",
        "reasons",
        "pitch",
    ]

    with open(path, "w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(build_priority_row(row) for row in sorted_rows)

    return len(sorted_rows)


def clean_specific_issue(reasons: object, pitch: object) -> str:
    first_reason = str(reasons or "").split(";")[0].strip()
    if first_reason:
        return first_reason.replace("_", " ")
    return str(pitch or "website improvements").strip()


def build_email_content(row: Dict[str, object]) -> Tuple[str, str]:
    score = as_float(row.get("score_0_100", 0))
    business_name = str(row.get("business_name", "your business")).strip() or "your business"
    specific_issue = clean_specific_issue(row.get("reasons", ""), row.get("pitch", ""))

    if score >= 75:
        tier = "Tier A"
    elif score >= 50:
        tier = "Tier B"
    else:
        tier = "Tier C"

    subject = f"Quick win idea for {business_name} ({tier})"
    body = (
        f"Hi {business_name} team,\n\n"
        "I took a quick look at your website and noticed one specific issue: "
        f"{specific_issue}.\n\n"
        f"You're currently in {tier} in our grading pass, and I can share a short plan to fix this quickly "
        "and improve inbound lead conversion.\n\n"
        "Would you like me to send over a 3-step recommendation?"
    )
    return subject, body


def build_form_message(row: Dict[str, object], city_name: str) -> str:
    business_name = str(row.get("business_name", "your business")).strip() or "your business"
    specific_issue = clean_specific_issue(row.get("reasons", ""), row.get("pitch", ""))
    category = str(row.get("category", DEFAULT_CATEGORY)).strip().lower() or DEFAULT_CATEGORY
    city_fragment = f" in {city_name}" if city_name else ""
    return (
        f"Hi {business_name} team — I ran a quick website audit for local {category} companies{city_fragment} and spotted one issue: "
        f"{specific_issue}. I can send a short 3-step fix plan to improve inbound leads. "
        "Would you like me to share it here?"
    )


def build_call_script(row: Dict[str, object], city_name: str) -> str:
    business_name = str(row.get("business_name", "your business")).strip() or "your business"
    specific_issue = clean_specific_issue(row.get("reasons", ""), row.get("pitch", ""))
    category = str(row.get("category", DEFAULT_CATEGORY)).strip().lower() or DEFAULT_CATEGORY
    city_fragment = f" in {city_name}" if city_name else ""
    return (
        f"Hi, this is Nathan — I did a quick website audit for local {category} companies{city_fragment}, including {business_name}, and noticed one issue: {specific_issue}. "
        "Could you point me to who handles your website, or the best email where I can send a short 3-step plan?"
    )


def populate_channel_messages(rows: List[Dict[str, object]], city_name: str) -> None:
    for row in rows:
        channel = str(row.get("outreach_channel", "none")).strip().lower()
        if channel == "none":
            row["form_message"] = ""
            row["call_script"] = ""
            continue

        row["form_message"] = build_form_message(row, city_name)
        row["call_script"] = build_call_script(row, city_name)


def write_outreach_sheet(path: str, prioritized_rows: List[Dict[str, object]], city_name: str) -> None:
    headers = [
        "business_name",
        "city",
        "phone",
        "primary_email",
        "emails",
        "has_contact_form",
        "contact_form_url",
        "outreach_channel",
        "form_message",
        "call_script",
        "url",
        "opportunity_score_0_100",
        "score_0_100",
        "reasons",
        "pitch",
        "email_subject",
        "email_body",
        "email_sent",
        "followup_1_sent",
        "followup_2_sent",
        "response",
        "status",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Outreach"
    sheet.append(headers)

    for row in prioritized_rows:
        email_subject, email_body = build_email_content(row)
        sheet.append(
            [
                row.get("business_name", ""),
                city_name,
                row.get("phone", ""),
                row.get("primary_email", ""),
                row.get("emails", ""),
                row.get("has_contact_form", False),
                row.get("contact_form_url", ""),
                row.get("outreach_channel", "none"),
                row.get("form_message", ""),
                row.get("call_script", ""),
                row.get("url", ""),
                row.get("opportunity_score_0_100", ""),
                row.get("score_0_100", ""),
                row.get("reasons", ""),
                row.get("pitch", ""),
                email_subject,
                email_body,
                "",
                "",
                "",
                "",
                "",
            ]
        )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for col_idx, _ in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in sheet[column_letter]:
            value_len = len(str(cell.value)) if cell.value is not None else 0
            if value_len > max_len:
                max_len = value_len
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 80)

    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Find and grade local businesses from OpenStreetMap.")
    parser.add_argument("--city", required=True, help="City name to search (e.g. 'Austin, TX').")
    parser.add_argument(
        "--category",
        default=DEFAULT_CATEGORY,
        help=f"Business category to search (allowed: {', '.join(sorted(CATEGORY_CONFIG.keys()))}; default: {DEFAULT_CATEGORY}).",
    )
    parser.add_argument("--limit", type=int, default=50, help="Maximum number of unique leads (default: 50).")
    parser.add_argument(
        "--radius_km",
        type=float,
        default=25,
        help="Search radius around city center in kilometers (default: 25).",
    )
    parser.add_argument(
        "--terms",
        default=",".join(DEFAULT_TERMS),
        help=(
            "Comma-separated terms for case-insensitive business name matching "
            "(default: plumber,plumbing,drain,sewer,rooter,pipe,water heater)."
        ),
    )
    parser.add_argument(
        "--overpass_timeout",
        type=int,
        default=180,
        help="Overpass query timeout in seconds (default: 180).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    category = str(args.category).strip().lower()
    if category not in CATEGORY_CONFIG:
        allowed = ", ".join(sorted(CATEGORY_CONFIG.keys()))
        raise ValueError(f"Unsupported category '{args.category}'. Allowed categories: {allowed}")

    category_config = CATEGORY_CONFIG[category]
    terms = parse_terms(args.terms, list(category_config.get("default_terms", DEFAULT_TERMS)))

    configured_sources = parse_lead_sources()
    osm_leads: List[Dict[str, object]] = []
    yelp_leads: List[Dict[str, object]] = []
    bing_leads: List[Dict[str, object]] = []

    if "osm" in configured_sources:
        osm_leads = discover_leads_osm(
            city=args.city,
            category=category,
            limit=max(args.limit * 2, args.limit),
            radius_km=args.radius_km,
            terms=terms,
            overpass_timeout=args.overpass_timeout,
        )

    if "yelp" in configured_sources:
        yelp_leads = discover_leads_yelp(
            city=args.city,
            limit=max(args.limit * 2, args.limit),
            radius_km=args.radius_km,
            category=category,
        )

    if "bing" in configured_sources:
        bing_leads = discover_leads_bing(
            city=args.city,
            limit=max(args.limit * 2, args.limit),
            radius_km=args.radius_km,
            category=category,
        )

    print(f"OSM leads: {len(osm_leads)}")
    yelp_key_present = bool(os.getenv("YELP_API_KEY", "").strip())
    bing_key_present = bool(os.getenv("BING_API_KEY", "").strip())
    if "yelp" not in configured_sources:
        print("Yelp leads: 0 (source disabled)")
    else:
        print(f"Yelp leads: {len(yelp_leads)}" + ("" if yelp_key_present else " (skipped if no key)"))
    if "bing" not in configured_sources:
        print("Bing leads: 0 (source disabled)")
    else:
        print(f"Bing leads: {len(bing_leads)}" + ("" if bing_key_present else " (skipped if no key)"))

    discovered_leads = [*osm_leads, *yelp_leads, *bing_leads]
    deduped = dedupe_leads(discovered_leads)
    ordered = sort_leads(deduped)
    leads = ordered[: args.limit]
    print(f"After dedupe: {len(deduped)} unique leads")

    write_raw_leads(RAW_OUTPUT_FILE, leads)

    leads_with_websites = [lead for lead in leads if str(lead.get("website_url") or lead.get("website") or "").strip()]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        graded_rows = list(
            executor.map(
                lambda lead: grade_website(str(lead.get("website_url") or lead.get("website") or "")),
                leads_with_websites,
            )
        )

    report_rows = [
        build_graded_row(lead, graded)
        for lead, graded in zip(leads_with_websites, graded_rows)
    ]
    city_name = str(args.city).split(",", maxsplit=1)[0].strip()
    populate_channel_messages(report_rows, city_name)
    write_graded_report(GRADED_OUTPUT_FILE, report_rows)
    priority_count = write_priority_leads(path=PRIORITY_OUTPUT_FILE, rows=report_rows)
    prioritized_rows = get_priority_rows(report_rows)
    write_outreach_sheet(OUTREACH_SHEET_OUTPUT_FILE, prioritized_rows, city_name)

    missing_phone_count = sum(1 for row in report_rows if not str(row.get("phone", "")).strip())

    print(f"Found {len(leads)} unique lead(s). Raw leads saved to {RAW_OUTPUT_FILE}.")
    print(
        f"Graded {len(report_rows)} lead(s) with websites. Report saved to {GRADED_OUTPUT_FILE}."
    )
    print(f"Created {PRIORITY_OUTPUT_FILE} with {priority_count} prioritized leads.")
    print(f"Created {OUTREACH_SHEET_OUTPUT_FILE} with {priority_count} prioritized leads.")
    print(f"Total graded websites: {len(report_rows)}")
    print(f"Included in {PRIORITY_OUTPUT_FILE}: {priority_count}")
    print(f"Missing phone: {missing_phone_count}")

    export_to_google_sheets(
        [
            (GRADED_OUTPUT_FILE, os.getenv("GOOGLE_SHEETS_TAB_REPORT", "report")),
            (PRIORITY_LEADS_OUTPUT_FILE, os.getenv("GOOGLE_SHEETS_TAB_PRIORITY", "priority_leads")),
            (RAW_OUTPUT_FILE, os.getenv("GOOGLE_SHEETS_TAB_LEADS_RAW", "leads_raw")),
        ]
    )


if __name__ == "__main__":
    main()
